#!/usr/bin/env python3
"""Export quiz records and monthly training records (Mar-Jul 2026) to Desktop Excel files."""

import sqlite3
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB_PATH = Path("/Users/hanying/peixun/app/data/quiz.db")
DESKTOP = Path("/Users/hanying/Desktop")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
ALT_FILL = PatternFill("solid", fgColor="D6E4F0")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def style_row(ws, row, cols, alt=False):
    fill = ALT_FILL if alt else PatternFill("solid", fgColor="FFFFFF")
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = Font(size=10)
        cell.alignment = LEFT
        cell.border = BORDER


def auto_width(ws, min_w=8, max_w=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                # CJK chars count as 2
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, length)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_w, max(min_w, max_len + 2))


def export_quiz_records(con):
    """抽问记录：每条答题记录一行（只含3-7月周期）"""
    cur = con.cursor()
    rows = cur.execute("""
        SELECT
            s.cycle_id,
            s.staff_id,
            s.staff_name,
            a.question_text,
            a.category,
            a.answer_text,
            a.score,
            a.level,
            a.summary,
            a.created_at
        FROM sessions s
        JOIN answers a ON s.id = a.session_id
        WHERE s.is_deleted = 0
          AND s.is_practice = 0
          AND (
            s.cycle_id LIKE 'cycle_2026-03%'
            OR s.cycle_id LIKE 'cycle_2026-04%'
            OR s.cycle_id LIKE 'cycle_2026-05%'
            OR s.cycle_id LIKE 'cycle_2026-06%'
            OR s.cycle_id LIKE 'cycle_2026-07%'
          )
        ORDER BY a.created_at
    """).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "3-7月抽问记录"
    ws.freeze_panes = "A2"

    headers = ["轮次", "工号", "姓名", "题目", "类别", "回答内容", "得分", "评级", "AI评价", "抽问时间"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    for i, row in enumerate(rows, 2):
        cycle_label = row[0].replace("cycle_", "").replace("-", "/") if row[0] else ""
        values = [cycle_label, row[1], row[2], row[3], row[4],
                  row[5], row[6], row[7], row[8], row[9]]
        for col, val in enumerate(values, 1):
            ws.cell(row=i, column=col, value=val)
        style_row(ws, i, len(headers), alt=(i % 2 == 0))

    auto_width(ws)
    # 题目列和回答列宽度固定大一些
    ws.column_dimensions["D"].width = 45
    ws.column_dimensions["F"].width = 50
    ws.column_dimensions["I"].width = 55
    ws.row_dimensions[1].height = 22

    out = DESKTOP / "3-7月抽问记录.xlsx"
    wb.save(out)
    print(f"✓ 抽问记录: {out}  ({len(rows)} 条答题记录)")
    return len(rows)


def export_training_records(con):
    """月度培训记录：每场培训一行（含出勤汇总）"""
    cur = con.cursor()
    plans = cur.execute("""
        SELECT
            m.id,
            m.year_month,
            m.shift_date,
            m.location,
            m.plan_type,
            m.leader_name,
            m.notes,
            m.completed_items,
            m.change_log
        FROM monthly_training_plans m
        WHERE m.year_month >= '2026-03' AND m.year_month <= '2026-07'
        ORDER BY m.shift_date
    """).fetchall()

    # 出勤汇总
    attendance = {}
    att_rows = cur.execute("""
        SELECT ta.plan_id, ta.staff_id, ta.checked_in, st.name as sname
        FROM training_attendance ta
        LEFT JOIN staff st ON ta.staff_id = st.id
        WHERE ta.plan_id IN (
            SELECT id FROM monthly_training_plans
            WHERE year_month >= '2026-03' AND year_month <= '2026-07'
        )
        ORDER BY ta.plan_id
    """).fetchall()

    for pid, sid, checked, sname in att_rows:
        if pid not in attendance:
            attendance[pid] = {"total": 0, "checked": 0, "names": []}
        attendance[pid]["total"] += 1
        if checked:
            attendance[pid]["checked"] += 1
            attendance[pid]["names"].append(sname or sid)

    wb = Workbook()
    ws = wb.active
    ws.title = "3-7月培训记录"
    ws.freeze_panes = "A2"

    headers = ["年月", "培训日期", "地点", "培训类型", "带教负责人",
               "备注", "已完成内容", "应到人数", "实到人数", "签到人员", "变更记录"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, 1, len(headers))

    for i, plan in enumerate(plans, 2):
        pid, year_month, shift_date, location, plan_type, leader, notes, completed_items_raw, change_log = plan
        att = attendance.get(pid, {})
        total = att.get("total", 0) or ""
        checked = att.get("checked", 0) or ""
        names = "、".join(att.get("names", [])) if att.get("names") else ""

        # Parse completed_items JSON
        try:
            items = json.loads(completed_items_raw or "[]")
            completed_str = "、".join(items) if items else ""
        except Exception:
            completed_str = completed_items_raw or ""

        values = [year_month, shift_date, location, plan_type, leader,
                  notes, completed_str, total, checked, names, change_log]
        for col, val in enumerate(values, 1):
            ws.cell(row=i, column=col, value=val)
        style_row(ws, i, len(headers), alt=(i % 2 == 0))

    auto_width(ws)
    ws.column_dimensions["G"].width = 40
    ws.column_dimensions["J"].width = 40
    ws.column_dimensions["K"].width = 55
    ws.row_dimensions[1].height = 22

    out = DESKTOP / "3-7月培训记录.xlsx"
    wb.save(out)
    print(f"✓ 培训记录: {out}  ({len(plans)} 场培训)")
    return len(plans)


def main():
    con = sqlite3.connect(DB_PATH)
    try:
        n_quiz = export_quiz_records(con)
        n_train = export_training_records(con)
        print(f"\n完成！已生成到桌面：")
        print(f"  3-7月抽问记录.xlsx  ({n_quiz} 条)")
        print(f"  3-7月培训记录.xlsx  ({n_train} 场)")
    finally:
        con.close()


if __name__ == "__main__":
    main()
